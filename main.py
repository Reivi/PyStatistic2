import math
import os

import openpyxl
from openpyxl.chart import Reference, BarChart
from openpyxl.styles import Font, Alignment

from PyQt5.QtGui import QDoubleValidator
import pyqtgraph as pg
from PyQt5 import QtCore

from PyQt5 import QtWidgets
from PyQt5.QtWidgets import QTableWidgetItem, QHeaderView, QFileDialog


import loadUi # Импорт файла loadUi.py
import sys


class ExampleApp(QtWidgets.QMainWindow, loadUi.Ui_MainWindow): # Класс в котором реализуется функционал приложения и наследутся от класса графического интерфейса (Графический интерфейс представлени в файле loadUi.py)
    def __init__(self):
        super(ExampleApp, self).__init__()
        self.setupUi(self)

        self.setWindowTitle("PyStatistic") # Название она приложения

        self.Spin_input_table.valueChanged.connect(self.change) # Соединение spin с функцией для изменения кол-ва строк в таблице

        self.Table_result.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Fixed) # Фиксированный размер ячеек в таюлице результатов
        self.Table_result.verticalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Fixed)# Фиксированный размер ячеек

        self.Table_input.setColumnCount(1) # кол-во столбцов в таблице ввода
        self.Table_input.setRowCount(int(self.Spin_input_table.text())) # кол-во строк в зависимости от числа в spin
        self.Table_input.setHorizontalHeaderLabels(["Даннные"]) # Название столбца
        self.Table_input.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Fixed) # Фиксированный размер ячеек
        self.Table_input.verticalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Fixed) # Фиксированный размер ячеек

        self.Table_statistic.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Fixed) # Фиксированный размер ячеек
        self.Table_statistic.verticalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Fixed) # Фиксированный размер ячеек


        self.window = pg.plot() # Создание графика 1
        self.Widget_Layout.addWidget(self.window) # Добавление графика в окно приложения
        self.window.setBackground('w')# Выбор цвета заднего фона для графика
        self.window.showGrid(y=True) # Линии сетки для графика
        self.window.setMouseEnabled(x=False, y=False) # Отключение, для графика, перемещения с помощь мыши

        self.window_Pi = pg.plot() # Создание графика 2
        self.Widget_Layout.addWidget(self.window_Pi) # Добавление графика в окно приложения
        self.window_Pi.setBackground('w')# Выбор цвета заднего фона для графика
        self.window_Pi.showGrid(y=True)# Линии сетки для графика
        self.window_Pi.setMouseEnabled(x=False, y=False)# Отключение, для графика, перемещения с помощь мыши

        self.Obr_button.clicked.connect(self.get_data)  # Соединение кнопки с функцией get_data
        self.clear_Button.clicked.connect(self.clear)
        self.action_Excel.triggered.connect(self.Export_to_Excel) # Соединение кнопки экспорта с функцией Export_to_Excel


        self.Name_graf.setPlaceholderText('Название проекта') # Подсказка для окна ввода
        self.Number_value.setPlaceholderText('Главное число')

        validator = QDoubleValidator(0.0, 1.0, 4)
        validator.setLocale(QtCore.QLocale("en_US"))
        self.Number_value.setValidator(validator)


    def clear(self):
        for i in range(self.Table_input.rowCount()):
            self.Table_input.setItem(i,0, QTableWidgetItem(None))
    def plot(self, axisX, axisY,Pi): # Функция для заполнения графиков
        self.window.clear() # очистка графика 1
        self.window.setLabel('left', 'Количество') # левая подпись графика 1
        self.window.setLabel('bottom', 'Интервалы')# нижняя подпись графика 1
        xdict = dict(enumerate(axisX)) # создание словаря нумерации для нижних значений графика

        ticks = [list(zip(range(7), (axisX)))] # создание списка для нижних значений графика
        xax = self.window.getAxis('bottom') # Присовение списка к низу графика
        xax.setTicks(ticks) # Вставка списка

        self.window.addItem(pg.BarGraphItem(x=list(xdict.keys()), height=axisY, width=0.6, brush='g')) # Стиль графика 1
        self.window.setTitle(self.Name_graf.text(), size="25pt", color="black")

        self.window_Pi.clear() # очистка графика 2
        self.window_Pi.setLabel('left', 'Pi') # левая подпись графика 2
        self.window_Pi.setLabel('bottom', 'Интервалы') # нижняя подпись графика 2
        xax = self.window_Pi.getAxis('bottom') # Присовение списка к низу графика
        xax.setTicks(ticks) # Вставка списка
        self.window_Pi.addItem(pg.BarGraphItem(x=list(xdict.keys()), height=Pi, width=0.6, brush='g')) # Стиль графика 2
        self.window_Pi.setTitle(self.Name_graf.text(), size="25pt", color="black")

    def Export_to_Excel(self): # Функиця экспортирования в Excel
        self.get_data() # Вызов функции получения данных с таблицы
        my_wb = openpyxl.Workbook() # создание экземпляра Excel файла
        my_sheet = my_wb.active
        c1 = my_sheet.cell(row=1, column=1) # Далее идет вставка данных из таблиц в Excel файл
        c1.value = "№ опыта"
        bold_font = Font(bold=True)
        c1.alignment = Alignment(horizontal='center')
        c1.font = bold_font
        c1 = my_sheet.cell(row=1, column=2)
        c1.value = "Данные"
        bold_font = Font(bold=True)
        c1.alignment = Alignment(horizontal='center')
        c1.font = bold_font
        c1 = my_sheet.cell(row=1, column=3)
        c1.value = "Измененные данные"
        bold_font = Font(bold=True)
        c1.alignment = Alignment(horizontal='center')
        c1.font = bold_font
        error_dialog = QtWidgets.QErrorMessage()
        try:
            for i in range(self.Table_input.rowCount()):
                c1 = my_sheet.cell(row=i + 2, column=2)
                c1.value = float(self.Table_input.item(i, 0).text())
                c1.alignment = Alignment(horizontal='center')
                c1 = my_sheet.cell(row=i + 2, column=1)
                c1.value = i + 1
                c1.alignment = Alignment(horizontal='center')
                if self.Number_value.text() != "":
                    c1 = my_sheet.cell(row=i + 2, column=3)
                    c1.value = float(self.Table_input.item(i, 1).text())
                    c1.alignment = Alignment(horizontal='center')
                    c1 = my_sheet.cell(row=i + 2, column=2)
                    c1.value = i + 1
                    c1.alignment = Alignment(horizontal='center')
        except:
            pass
        c1 = my_sheet.cell(row=1, column=5)
        c1.value = "Интервал"
        c1.font = bold_font
        c1 = my_sheet.cell(row=1, column=6)
        c1.value = "Количество"
        c1.font = bold_font
        c1 = my_sheet.cell(row=1, column=7)
        c1.value = "Pi"
        c1.font = bold_font
        c1.alignment = Alignment(horizontal='center')

        for i in range(7):
            c1 = my_sheet.cell(row=i + 2, column=5)
            try:
                c1.value = self.Table_result.item(i, 0).text()
            except:
                error_dialog.showMessage('Данные не обработаны!')
            c1.alignment = Alignment(horizontal='center')
            c1 = my_sheet.cell(row=i + 2, column=6)
            c1.value = float(self.Table_result.item(i, 1).text())
            c1.alignment = Alignment(horizontal='center')
            c1 = my_sheet.cell(row=i + 2, column=7)
            c1.value = float(self.Table_result.item(i, 2).text())
            c1.alignment = Alignment(horizontal='center')

        c1 = my_sheet.cell(row=1, column=10)
        c1.value = "Параметр"
        c1.font = bold_font
        c1 = my_sheet.cell(row=1, column=9)
        c1.value = "Название"
        c1.font = bold_font
        c1 = my_sheet.cell(row=2, column=9)
        c1.value = "Дисперсия"
        c1.font = bold_font
        c1 = my_sheet.cell(row=3, column=9)
        c1.value = "Сред. квад. окл."
        c1.font = bold_font
        c1 = my_sheet.cell(row=4, column=9)
        c1.value = "Среднее знач."
        c1.font = bold_font
        c1 = my_sheet.cell(row=5, column=9)
        c1.value = "Коэф. вариац."
        c1.font = bold_font
        c1 = my_sheet.cell(row=6, column=9)
        c1.value = "Мин."
        c1.font = bold_font
        c1 = my_sheet.cell(row=7, column=9)
        c1.value = "Макс."
        c1.font = bold_font

        for i in range(6):
            c1 = my_sheet.cell(row=i + 2, column=10)
            c1.value = self.Table_statistic.item(i, 1).text()
            c1.alignment = Alignment(horizontal='center')

        chart1 = BarChart()
        # установим тип - `вертикальные столбцы`
        chart1.type = "col"
        # установим стиль диаграммы (цветовая схема)
        chart1.style = 8

        chart1.y_axis.title = 'Количество'
        chart1.y_axis.delete = False
        chart1.x_axis.title = 'Интервалы'
        chart1.x_axis.delete = False
        chart1.width = 15
        chart1.height = 11
        chart1.title = self.Name_graf.text()
        # выберем 2 столбца с данными для оси `y`
        data = Reference(my_sheet, min_col=6, max_col=6, min_row=1, max_row=8)
        # теперь выберем категорию для оси `x`
        categor = Reference(my_sheet, min_col=5, max_col=5, min_row=2, max_row=8)
        # добавляем данные в объект диаграммы
        chart1.add_data(data, titles_from_data=True)
        # установим метки на объект диаграммы
        chart1.set_categories(categor)
        my_sheet.add_chart(chart1, "L2")

        chart2 = BarChart()
        # установим тип - `вертикальные столбцы`
        chart2.type = "col"
        # установим стиль диаграммы (цветовая схема)
        chart2.style = 8

        chart2.y_axis.title = 'Pi'
        chart2.y_axis.delete = False
        chart2.x_axis.title = 'Интервалы'
        chart2.x_axis.delete = False
        chart2.width = 15
        chart2.height = 11
        chart2.title = self.Name_graf.text()
        # выберем 2 столбца с данными для оси `y`
        data = Reference(my_sheet, min_col=7, max_col=7, min_row=1, max_row=8)
        # теперь выберем категорию для оси `x`
        categor = Reference(my_sheet, min_col=5, max_col=5, min_row=2, max_row=8)
        # добавляем данные в объект диаграммы
        chart2.add_data(data, titles_from_data=True)
        # установим метки на объект диаграммы
        chart2.set_categories(categor)
        my_sheet.add_chart(chart2, "L24")

        if self.Name_graf.text() != "": # Сохранение файла в форамате xlsx
            save_name = self.Name_graf.text() + ".xlsx"
        else:
            for i in range(20):
                if not os.path.exists("Data_Excel_" + str(i) + ".xlsx"):
                    save_name = "Data_Excel_" + str(i) + ".xlsx"
                    break
                else:
                    continue
                    break

        fileName, _ = QFileDialog.getSaveFileName(self, "Экспорт данных в Excel", save_name,
                                                  "Excel Files (*.xlsx);;All Files (*)")
        if fileName:
            my_wb.save(fileName)



    def date_statistic_set(self, variance, sred_kvad_otkl, sr_snaz, koev_variazii, maxX, minX): # Фунеция вставки обработанных данных в таблицу со статистиков
        self.Table_statistic.setRowCount(6) # установка 6 строк
        self.Table_statistic.setColumnCount(2) # установка 2 столбцов
        self.Table_statistic.setHorizontalHeaderLabels(["Параметр", "Значение"]) # Название столбцов
        self.Table_statistic.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch) # Установка размера ячеек
        self.Table_statistic.setItem(0, 0, QTableWidgetItem("Дисперсия")) #Встака в левую часть таблицы
        self.Table_statistic.setItem(1, 0, QTableWidgetItem("Среднее квад. откл."))
        self.Table_statistic.setItem(2, 0, QTableWidgetItem("Среднее значение"))
        self.Table_statistic.setItem(3, 0, QTableWidgetItem("Коэффициент вариации"))
        self.Table_statistic.setItem(4, 0, QTableWidgetItem("Минимальное"))
        self.Table_statistic.setItem(5, 0, QTableWidgetItem("Максимальное"))

        self.Table_statistic.setItem(0, 1, QTableWidgetItem(str(round(variance, 4)))) # Вставка в правую часть таблицы
        self.Table_statistic.setItem(1, 1, QTableWidgetItem(str(round(sred_kvad_otkl, 4))))
        self.Table_statistic.setItem(2, 1, QTableWidgetItem(str(round(sr_snaz, 4))))
        self.Table_statistic.setItem(3, 1, QTableWidgetItem(str(round(koev_variazii, 4))))
        self.Table_statistic.setItem(4, 1, QTableWidgetItem(str(round(minX, 4))))
        self.Table_statistic.setItem(5, 1, QTableWidgetItem(str(round(maxX, 4))))

    def date_set(self, kol_interval, interval, Pi): # Вставка обработанных данных в таблицу с интервалами

        self.Table_result.setRowCount(kol_interval) # Установка кол-ва строк
        self.Table_result.setColumnCount(3) # установка кол-ва столбцов
        self.Table_result.setHorizontalHeaderLabels(["Интервал", "Количество", "Pi"]) # название столбцов

        spam = [1 * 3 for i in range(kol_interval)] # создание списка для передачи в функцию plot
        tatl = [1 * 3 for i in range(kol_interval)] # создание списка для передачи в функцию plot
        for i in range(kol_interval): # Цикл для вставки данных в таблицу и заполнения списков для передачи в plot
            spam[i] = interval[i][2] # Кол-во входящих в интервал чисел
            tatl[i] = QTableWidgetItem( # Интервалы
                str("{}-{}".format(round(interval[i][0], 1),round(interval[i][1], 1)))).text()
            self.Table_result.setItem(i, 0, QTableWidgetItem(str(tatl[i]))) # Вставка Интервалов
            self.Table_result.setItem(i, 1, QTableWidgetItem(str(interval[i][2]))) # вставка кол-ва входящих в интервал чисел
            self.Table_result.setItem(i, 2, QTableWidgetItem(str(round(Pi[i], 1)))) # Вствка p[i]
        self.plot(tatl, spam, Pi) # Вызов функции plot

    def change(self): # Функция для динамического изменения кол-ва строк с таблице ввода с помощью spin
        self.Table_input.setRowCount(int(self.Spin_input_table.text()))

    def get_data(self): # Функция для получения данных с таблицы ввода с дальнейшей их обработкой

        average = 0 #Кол-во чисел
        summ = 0 # Сумма чисел
        maxX = -sys.maxsize - 1 # Максимум
        minX = sys.maxsize # Минимум
        Data = [] #Список всех чисел

        if self.Number_value.text() =="":
            column_number = 0
            self.Table_input.setColumnCount(1)
        else:
            self.Table_input.setColumnCount(2)
            column_number = 1
            for i in range(self.Table_input.rowCount()):
                self.Table_input.setItem(i, 1, QTableWidgetItem(None))

                try:
                    val = float(self.Number_value.text()) - float(self.Table_input.item(i, 0).text())
                    self.Table_input.setItem(i, 1, QTableWidgetItem(str(round(val, 4))))
                except:
                    pass

        for i in range(self.Table_input.rowCount()): # Цикл для заполнения списка и выщитывания значений summ и average
            if self.Table_input.item(i, column_number) is not None: # условние если ячейка не пустая, то...
                try: # Игнорирование исключений, связанных с некорректным вводом данных
                    Data.append(float(self.Table_input.item(i, column_number).text())) # Добавление в список текущего числа
                    summ = float(self.Table_input.item(i, column_number).text()) + summ # Прибавление к сумме текущего числа
                    average += 1
                    if maxX < float(self.Table_input.item(i, column_number).text()): # поиск максимума
                        maxX = float(self.Table_input.item(i, column_number).text())
                    if minX > float(self.Table_input.item(i, column_number).text()): # поиск минимума
                        minX = float(self.Table_input.item(i, column_number).text())

                except:# Игнорирование исключений, связанных с некорректным вводом данных
                    pass# Игнорирование исключений, связанных с некорректным вводом данных

        if average != 0: # Условие если кол-во чисел не равно 0, то...
            kol_interval = 7  # int(1 + abs(log(average, 2))) # В комментарии формула для динамического изменения кол-ва интервала, фиксированный интервал = 7
            sr_snaz = summ / len(Data) # Ширина значений

            deviations = [(x - sr_snaz) ** 2 for x in Data] # спсок

            variance = sum(deviations) / len(Data) # Дисперсия
            sred_kvad_otkl = math.sqrt(variance) # квадрат дисперции
            koev_variazii = sred_kvad_otkl / sr_snaz # Коэф. вариации

            shir_interval = (maxX - minX) / kol_interval # Ширина интервала
            interval = [[0] * 3 for i in range(kol_interval)] # инициализация списка для интервала
            interval[0][0] = minX # начало первого интервала
            interval[0][1] = (interval[0][0] + shir_interval) # Конец первого интервала
            numbers_in_interval = 0 # номер интервала

            for i in range(1, kol_interval): # цикл присвоения интервалов списку
                interval[i][0] = interval[i - 1][1] # начальное число интервала
                interval[i][1] = (interval[i][0] + shir_interval) # конечное число интервала

            for b in range(kol_interval): # Циклы для проверки входит ли число в интервал, если входит, то numbers_in_interval + 1
                for i in range(self.Table_input.rowCount()): # цикл выбора всех значений в таблице ввода
                    try: # Игнорирование исключений
                        if interval[b][0] <= float(self.Table_input.item(i, column_number).text()) <= round(interval[b][1], 10): # условие воходит ли в интервал
                            numbers_in_interval += 1 # Кол-во чисел в интервале + 1
                    except:# Игнорирование исключений
                        pass# Игнорирование исключений
                interval[b][2] = numbers_in_interval # Присвоение списку кол-ва входящих в данный интервал чисел
                numbers_in_interval = 0 # обнуление переменной

            Pi = [1 * 1 for i in range(kol_interval)] # инициализация списка для P[i]

            for i in range(kol_interval): # цикл для поиска p[i]
                Pi[i] = float((1 / len(Data)) * interval[i][2]) # Поиск p[i] в интервале

            self.date_set(kol_interval, interval, Pi) # вызов функции data_set для вставки в таблицу интервалов
            self.date_statistic_set(variance, sred_kvad_otkl, sr_snaz, koev_variazii, maxX, minX) # вызов функции date_statistic_set для вставки в таблицу статистики


app = QtWidgets.QApplication([])
application = ExampleApp()
application.showMaximized()
application.show()

sys.exit(app.exec())
